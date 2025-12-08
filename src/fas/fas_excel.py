import openpyxl
from typing import Dict, Any

def extract_excel_data(file_path: str) -> Dict[str, Any]:
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        sheet_names = workbook.sheetnames

        metadata = {
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
        }

        # Workbook document properties
        props = workbook.properties
        metadata.update({
            "creator": props.creator,
            "last_modified_by": props.lastModifiedBy,
            "created": str(props.created) if props.created else None,
            "modified": str(props.modified) if props.modified else None,
            "title": props.title,
            "subject": props.subject,
            "keywords": props.keywords,
            "category": props.category,
            "description": props.description
        })

        # Per sheet stats
        sheet_stats = {}
        total_formulas = 0
        total_charts = 0
        for sheet in workbook.worksheets:
            formulas = sum(
                1 for row in sheet.iter_rows() 
                for cell in row 
                if cell.value and str(cell.value).startswith("=")
            )
            total_formulas += formulas

            # Detect charts
            charts = len(sheet._charts) if hasattr(sheet, "_charts") else 0
            total_charts += charts

            sheet_stats[sheet.title] = {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "formulas": formulas,
                "merged_cells": len(sheet.merged_cells.ranges),
                "charts": charts
            }

        metadata["sheet_stats"] = sheet_stats


        # Infer key skills ----------------------------------------------------------------------------
        key_skills = []

        # Analytical skills: multiple sheets or structured sheets
        if len(sheet_names) > 1 or any(sheet_stats.values()):
            key_skills.append("Analytical Skills")

        # Excel proficiency: any formulas
        if total_formulas > 0:
            key_skills.append("Excel Proficiency")

        # Data visualization: any charts
        if total_charts > 0:
            key_skills.append("Data Visualization")

        # Financial modeling: detect common finance functions
        finance_functions = ["NPV", "IRR", "PMT", "FV"]
        finance_detected = False
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        for func in finance_functions:
                            if func in cell.value.upper():
                                finance_detected = True
                                break
        if finance_detected:
            key_skills.append("Financial Modeling")

        metadata["key_skills"] = key_skills

        workbook.close()
        return metadata

    except Exception as e:
        return {"error": str(e)}
