import os
import openpyxl
import pytest
from src.fas.fas_excel import extract_excel_data

# Test directories and files
TESTDATA_DIR = os.path.join("tests", "testdata", "test_fas")
TEST_FILE_PATH = os.path.join(TESTDATA_DIR, "fas_excel_test.xlsx")
INVALID_FILE_PATH = os.path.join(TESTDATA_DIR, "invalid_excel.xlsx")


def create_excel_test_file():
    os.makedirs(TESTDATA_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = 10
    ws1["B1"] = 20
    ws1["C1"] = "=SUM(A1,B1)"  # formula
    ws1.merge_cells("A2:B2")    # merged cell

    ws2 = wb.create_sheet("Data")
    ws2["A1"] = 123
    ws2["A2"] = 456

    ws3 = wb.create_sheet("HelloWorld")
    ws3["A1"] = "Hello"
    ws3["B1"] = "World"

    # Workbook metadata
    wb.properties.creator = "Toby"
    wb.properties.lastModifiedBy = "Nguyen"
    wb.properties.title = "Test Workbook"
    wb.properties.subject = "Testing"
    wb.properties.keywords = "Excel,Test"
    wb.properties.category = "Test Category"
    wb.properties.description = "This is a test workbook"

    wb.save(TEST_FILE_PATH)
    wb.close()

    # Re-open workbook for formulas to register
    wb = openpyxl.load_workbook(TEST_FILE_PATH)
    wb.save(TEST_FILE_PATH)
    wb.close()

    return TEST_FILE_PATH


def create_invalid_excel_file():
    os.makedirs(TESTDATA_DIR, exist_ok=True)
    with open(INVALID_FILE_PATH, "w") as f:
        f.write("Not a real Excel file")
    return INVALID_FILE_PATH


# Start of tests------------------------------------------------------
class TestFasExcel:

    def test_sheet_count_and_names(self):
        file_path = create_excel_test_file()
        metadata = extract_excel_data(file_path)
        assert metadata["sheet_count"] == 3
        assert set(metadata["sheet_names"]) == {"Summary", "Data", "HelloWorld"}


    def test_workbook_metadata(self):
        metadata = extract_excel_data(TEST_FILE_PATH)
        assert metadata["creator"] == "Toby"
        assert metadata["last_modified_by"] == "Nguyen"
        assert metadata["title"] == "Test Workbook"
        assert metadata["subject"] == "Testing"
        assert metadata["keywords"] == "Excel,Test"
        assert metadata["category"] == "Test Category"
        assert metadata["description"] == "This is a test workbook"


    def test_sheet_stats(self):
        metadata = extract_excel_data(TEST_FILE_PATH)
        stats = metadata["sheet_stats"]

        # Summary sheet
        summary = stats["Summary"]
        assert summary["max_row"] >= 2
        assert summary["max_column"] >= 3
        assert summary["formulas"] == 1
        assert summary["merged_cells"] == 1

        # Data sheet
        data = stats["Data"]
        assert data["max_row"] >= 2
        assert data["max_column"] >= 1
        assert data["formulas"] == 0
        assert data["merged_cells"] == 0

        # HelloWorld sheet
        hello = stats["HelloWorld"]
        assert hello["max_row"] >= 1
        assert hello["max_column"] >= 2


    def test_invalid_excel_file(self):
        bad_file = create_invalid_excel_file()
        metadata = extract_excel_data(bad_file)
        assert "error" in metadata
        assert isinstance(metadata["error"], str)
